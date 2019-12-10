import os
import datetime
import json
import csv
import string
from collections import OrderedDict
import openpyxl

"""
Written for the legends of ESP (https://sepush.co.za/)
Hopefully this helps to keep on top of the wonderful joy
that Eskom call LoadShedding.
"""

def clean_string(string_to_clean):
    valid_chars = "-_.() %s%s" % (string.ascii_letters, string.digits)
    return ''.join(c for c in string_to_clean if c in valid_chars)

class SheetReader:
    def __init__(self, sheetname):
        self.sch_range = [16, 111]  # Hidden bit of magic in the files
        self.workbook = openpyxl.load_workbook(sheetname)


    def load_areas(self):
        self.areas = []
        sheet = self.workbook.get_sheet_by_name('SP_List')  # Hidden sheet behind the dropdown
        row = 2
        while True:
            if sheet.cell(row=row, column=1).value:
                area = OrderedDict({'block': sheet.cell(row=row, column=7).value, 'municipality': sheet.cell(row=row, column=2).value, 'area_name': sheet.cell(row=row, column=4).value})
                self.areas.append(area)
            else:
                break
            row += 1

    def load_schedule(self):
        self.schedules = []
        sheet = self.workbook.get_sheet_by_name('Schedule (4H)')  # Main schedule sheet
        self.province = sheet['A6'].value
        start_time = ''
        end_time = ''
        for row in range(self.sch_range[0], self.sch_range[1]):
            if sheet.cell(row=row,column=3).value == 1:  # Merged cells
                start_time = sheet.cell(row=row,column=1).value.strftime("%H:%M:%S")
                end_time = sheet.cell(row=row,column=2).value.strftime("%H:%M:%S")
            dblock = []
            for col in range(4,35):  # Day / Block mapping
                dblock.append([col-3, sheet.cell(row=row,column=col).value])
            sch = OrderedDict({'start_time': start_time,
                   'end_time': end_time,
                   'stage': sheet.cell(row=row,column=3).value,
                   'day-block': dblock })
            self.schedules.append(sch)

    def match_areas_with_schedules(self):
        self.sch_area = []
        for area in self.areas:
            block = area.get('block')
            for sch in self.schedules:
                sa = OrderedDict({'stage': sch.get('stage'),
                      'start_time': sch.get('start_time'),
                      'end_time': sch.get('end_time'),
                      'block': block,
                      'municipality': area.get('municipality'),
                      'area_name': area.get('area_name')})
                for dblock in sch.get('day-block'):
                    if dblock[1] == block:
                        sa[dblock[0]] = 'Y'
                    else:
                        sa[dblock[0]] = 'N'
                self.sch_area.append(sa)


    def write_files(self):
        outputdir = os.path.join(os.getcwd(), 'output', self.province)
        if not os.path.exists(outputdir):
            os.makedirs(outputdir)
        for area in self.areas:
            print(area, area.get('municipality'))
            with open(os.path.join(outputdir, clean_string(area.get('municipality')) + ' ' + clean_string(area.get('area_name')) + '.csv'), 'w') as csvfile:
                fieldnames = list(self.sch_area[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                for row in self.sch_area:
                    if row.get('area_name') == area.get('area_name'):
                        writer.writerow(row)

if __name__ == '__main__':
    srcdir = os.path.join(os.getcwd(), 'schedules')  # from http://www.eskom.co.za/Pages/LS_schedules.aspx
    for file in os.listdir(srcdir):
        if file.endswith('.xlsx'):
            s = SheetReader(os.path.join(srcdir, file))
            s.load_areas()
            s.load_schedule()
            s.match_areas_with_schedules()
            s.write_files()